import tkinter as tk
from tkinter import messagebox
from Connection import Control 
import openpyxl
from openpyxl import Workbook


class Admin_View_Enrollment_Class:
    def __init__(self, root:tk.Tk,admin_id, admin_name,role):
        self.admin_id = admin_id
        self.admin_name = admin_name
        self.role = role
        
        self.win = root
        self.db = Control()
        self.win.title("Admin View Enrollment Courses")
        self.win.geometry("500x400")
        self.win.iconbitmap("images/icon_image.ico")
        self.win.resizable(False,False)
        self.win.configure(bg="#208dc4")
        self.win.resizable(False, False)
        
        outer_frame = tk.Frame(self.win)
        outer_frame.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(outer_frame, bg="#208dc4")
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(outer_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self.frame = tk.Frame(canvas, bg="#208dc4")
        canvas.create_window((0, 0), window=self.frame, anchor="nw")
        

        exportbutton = tk.Button(self.frame, text="Export to Excel", font=12, bg="green", fg="white", command=self.export_excel)
        exportbutton.grid(row=0, column=0, pady=20)
        button1 = tk.Button(self.frame, text= "Back", bg="red", fg="white",font=12,command=lambda: self.back(self.admin_id, self.admin_name, self.role))
        button1.grid(row=0 , column=1,pady=20)
       
        label1=tk.Label(self.frame, text="Username", font=("Arial", 12, "bold"), bg="#208dc4", fg="white")
        label1.grid(row=1, column=0, padx=10, pady=10)
        
        label2=tk.Label(self.frame, text="Name", font=("Arial", 12, "bold"), bg="#208dc4", fg="white")
        label2.grid(row=1, column=1, padx=10, pady=10)
        
        label3=tk.Label(self.frame, text="Course Name", font=("Arial", 12, "bold"), bg="#208dc4", fg="white")
        label3.grid(row=1, column=2, padx=10, pady=10)
        
        label4=tk.Label(self.frame, text="Enrollment Date", font=("Arial", 12, "bold"), bg="#208dc4", fg="white")
        label4.grid(row=1, column=3, padx=10, pady=10)
        self.Get_Enrollment()

    
        
        
    def back(self,admin_id, admin_name,role):
        self.win.destroy()
        import Admin_Deshboard
        
        win2 = tk.Tk()
        Admin_Deshboard.Admin(win2,admin_id, admin_name,role)
        

    def Get_Enrollment(self):
        
        try:
            self.db.cursor.execute("SELECT u.username, u.name , c.course_name, e.enrollment_date from tblEnrollments e join tblUsers u on e.users_id = u.id join tblCourses c ON e.course_id = c.id")
            self.View_Enroll = self.db.cursor.fetchall()
            
            for i, data in enumerate(self.View_Enroll):
                id, std_name, course_name, enroll_date = data
                
            
                label5=tk.Label(self.frame, bg="#208dc4", text=id , font=12)
                label5.grid(row=i+2, column=0, pady=10 )
                
                label6=tk.Label(self.frame, bg="#208dc4", text=std_name , font=12)
                label6.grid(row=i+2, column=1, pady=10 )
                
                label7=tk.Label(self.frame, bg="#208dc4", text=course_name, font=12)
                label7.grid(row=i+2, column=2, pady=10 )
                
                label8=tk.Label(self.frame, bg="#208dc4", text=enroll_date , font=12)
                label8.grid(row=i+2, column=3, pady=10 )
           
        
        except Exception as e:
            messagebox.showerror("error",f"Something Wrong\n{e}" )
         

    def export_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Enrollments Report"

            headers = ["Username", "Name", "Course Name", "Enrollment Date"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

        
            for row, row_data in enumerate(self.View_Enroll, start=2):
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=row, column=col, value=value)

            file_name = "Enrollments_Report.xlsx"
            wb.save(file_name)
            messagebox.showinfo("Success", f"Excel report saved as '{file_name}'")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{e}")
            
